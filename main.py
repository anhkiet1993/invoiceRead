import os
import xml.etree.ElementTree as ET
import pandas as pd
from openpyxl import Workbook, load_workbook
from typing import List
import tkinter as tk
from tkinter import filedialog

class ProductItem:
    ProductID: str = ""
    ProductName: str = ""
    Unit: str = ""
    Quantity: str = ""
    UnitPrice: str = ""
    Discount: str = ""
    TotalUnitPrice: str = ""
    TaxRate: str = ""

class InvoiceData:
    ID: str = ""
    number: str = ""
    DateHD: str = ""
    DateCT: str = ""
    CTNo: str = ""
    CustomerID: str = ""
    BuyerName: str = ""
    CompanyName: str = ""
    Address: str = ""
    TaxCode: str = ""
    Reason: str = ""
    Products: List[ProductItem] = []
    TotalBeforeTax: str = ""
    TaxAmount: str = ""
    TotalPayment: str = ""

def get_xml_files(path):
    xml_files = []
    for root, dirs, files in os.walk(path):
        for file in files:
            if file.endswith(".xml"):
                xml_files.append(os.path.join(root, file))
    return xml_files

def create_excel_with_headers(file_name, sheet_name, headers):
    """Create a new Excel file with header row."""
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    ws.append(headers)
    wb.save(file_name)
    # print(f"Created '{file_name}' with headers: {headers}")


def append_row(file_name, sheet_name, row_data):
    """Append a single row of data to an existing Excel file."""
    wb = load_workbook(file_name)
    ws = wb[sheet_name]
    ws.append(row_data)
    wb.save(file_name)
    # print(f"Appended row: {row_data}")

def parse_xml(file):

    # append_row("InvoiceData.xlsx", "Data", ['Alice Smith', 'Home Appliance', 1234])

    invoice = InvoiceData()

    try:
        tree = ET.parse(file)
        root = tree.getroot()

        for item in root.findall('./DLHDon/TTChung'):
            for subItem in item:
                if (subItem.tag == "KHHDon"):
                    invoice.ID = subItem.text
                elif (subItem.tag == "SHDon"):
                    invoice.number = subItem.text
                elif (subItem.tag == "NLap"):
                    invoice.DateHD = subItem.text
                    invoice.DateCT = subItem.text

        for item in root.findall('./DLHDon/NDHDon/NMua'):
            for subItem in item:
                if (subItem.tag == "Ten"):
                    invoice.BuyerName = subItem.text
                elif (subItem.tag == "DChi"):
                    invoice.Address = subItem.text
                elif (subItem.tag == "MST"):
                    invoice.TaxCode = subItem.text
                elif (subItem.tag == "MKHang"):
                    invoice.CustomerID = subItem.text
        
        for item in root.findall('./DLHDon/NDHDon/TToan'):
            for subItem in item:
                if (subItem.tag == "TgTCThue"):
                    invoice.TotalBeforeTax = subItem.text
                elif (subItem.tag == "TgTThue"):
                    invoice.TaxAmount = subItem.text
                elif (subItem.tag == "TgTTTBSo"):
                    invoice.TotalPayment = subItem.text

        itemList = root.findall('./DLHDon/NDHDon/DSHHDVu/HHDVu')
        totalTtems = len(itemList)
        for index, item in enumerate(itemList):
            product = ProductItem()
            for subElement in item:
                if subElement.tag == "MHHDVu":
                    product.ProductID = subElement.text
                elif subElement.tag == "THHDVu":
                    product.ProductName = subElement.text
                elif subElement.tag == "DVTinh":
                    product.Unit = subElement.text
                elif subElement.tag == "SLuong":
                    product.Quantity = subElement.text
                elif subElement.tag == "DGia":
                    product.UnitPrice = subElement.text
                elif subElement.tag == "ThTien":
                    product.TotalUnitPrice = subElement.text
                elif subElement.tag == "STCKhau":
                    product.Discount = subElement.text
                elif subElement.tag == "TSuat":
                    product.TaxRate = subElement.text

            
            
            if index == totalTtems - 1:
                append_row("InvoiceData.xlsx", "Data", 
                            [invoice.ID, invoice.number, invoice.DateHD, invoice.DateCT, invoice.CTNo, invoice.CustomerID, 
                            invoice.BuyerName, invoice.CompanyName, invoice.Address, invoice.TaxCode, invoice.Reason, 
                            product.ProductID, product.ProductName, product.Unit, product.Quantity, product.UnitPrice,
                            product.TotalUnitPrice, product.Discount, product.TaxRate, invoice.TotalBeforeTax, invoice.TaxAmount, invoice.TotalPayment])
            else:
                append_row("InvoiceData.xlsx", "Data", 
                            [invoice.ID, invoice.number, invoice.DateHD, invoice.DateCT, invoice.CTNo, invoice.CustomerID, 
                            invoice.BuyerName, invoice.CompanyName, invoice.Address, invoice.TaxCode, invoice.Reason, 
                            product.ProductID, product.ProductName, product.Unit, product.Quantity, product.UnitPrice,
                            product.TotalUnitPrice, product.Discount, product.TaxRate])


        # Add more parsing logic here as needed
    except ET.ParseError as e:
        print(f"Error parsing {file}: {e}")

def select_folder():
    root = tk.Tk()
    root.withdraw()
    
    folder_path = filedialog.askdirectory(
        title="Select a Folder for Processing"
    )
    
    # 3. Check if a folder was selected or if the user clicked 'Cancel'
    if folder_path:
        print(f"Successfully selected folder: {folder_path}")
        return folder_path
    else:
        print("No folder was selected. The operation was canceled.")
        return None

def main():
    path = select_folder()

    if path:
        xml_files = get_xml_files(path)
        if not xml_files:
            print("No XML files found.")
            return

        # 1. Define headers and create the file
        headers = ['Ky Hieu', 'So hoa don', 'Ngay HD', 'Ngay CT', 'So CT', 'Ma KH', 'To ten nguoi mua', 'Ten don vi', 'Dia chi',
                    'Ma so thue', 'Ly do', 'Ma HH', 'Ten hang hoa dich vu', 'Don vi tinh', 'So luong', 'Don gia', 'Thanh tien', 
                    'Chiet khau', 'Thue suat', 'Tong tien truoc thue', 'Tien thue', 'Tong tien thanh toan']

        create_excel_with_headers("InvoiceData.xlsx", "Data", headers)

        for xml_file in xml_files:
            print(f"Parsing file: {xml_file}")
            parse_xml(xml_file)

if __name__ == "__main__":
    main()